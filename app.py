from flask import Flask, render_template, request, redirect, session
import openpyxl, os
from datetime import datetime
from routes.auth import auth_bp

app = Flask(__name__)
app.secret_key = 'secretkey'

# ✅ Register the Blueprint
app.register_blueprint(auth_bp, url_prefix='/')

# Project folder paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')  # For login/registration
EXTERNAL_DATA_DIR = "D:/AppData"           # All form/module Excel files

# User Excel files
USERS_FILE = os.path.join(DATA_DIR, 'users.xlsx')
USER_DETAILS_FILE = os.path.join(DATA_DIR, 'user_details.xlsx')

# Ensure login Excel files exist
def init_user_excel(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        if 'users' in file_path:
            ws.append(["Employee Code", "Name", "Designation", "Department", "Username", "Password"])
        elif 'user_details' in file_path:
            ws.append(["Employee Code", "Name", "Designation", "Department"])
        wb.save(file_path)

init_user_excel(USERS_FILE)
init_user_excel(USER_DETAILS_FILE)

# Auto-create Excel for forms
def ensure_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_path)

# ---------------- ROUTES ----------------

@app.route('/')
def home():
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4] == username and row[5] == password:
                session['name'] = row[1]
                return redirect('/dashboard')
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        emp_code = request.form['emp_code']
        username = request.form['username']
        password = request.form['password']
        confirm = request.form['confirm']

        if password != confirm or len(password) < 8:
            return render_template('register.html', error="Invalid or mismatched password")

        wb_details = openpyxl.load_workbook(USER_DETAILS_FILE)
        ws_details = wb_details.active
        found = False
        for row in ws_details.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == emp_code:
                name, designation, department = row[1:4]
                found = True
                break
        if not found:
            return render_template('register.html', error="Employee code not found")

        wb_users = openpyxl.load_workbook(USERS_FILE)
        ws_users = wb_users.active

        # ✅ Prevent duplicate by Employee Code
        for row in ws_users.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == emp_code:
                return render_template('register.html', error="This employee has already registered.")

        ws_users.append([emp_code, name, designation, department, username, password])
        wb_users.save(USERS_FILE)
        return redirect('/login')
    return render_template('register.html')

@app.route('/dashboard')
def dashboard():
    if 'name' not in session:
        return redirect('/login')
    return render_template('dashboard.html', name=session['name'])

@app.route('/defects', methods=['GET', 'POST'])
def defects():
    if 'name' not in session:
        return redirect('/login')

    file_path = os.path.join(EXTERNAL_DATA_DIR, "defects_data", "defects.xlsx")
    headers = ["Sr. No.", "Date", "Defect Stage", "Description", "Qty", "FG Part No.",
               "Project", "Line", "Supplier", "Operator", "Inspector", "Production Qty"]
    ensure_excel_file(file_path, headers)

    if request.method == 'POST':
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        sr_no = ws.max_row
        ws.append([
            sr_no,
            datetime.now().strftime('%Y-%m-%d'),
            request.form['stage'],
            request.form['description'],
            request.form['qty'],
            request.form['fg_part'],
            request.form['project'],
            request.form['line'],
            request.form['supplier'],
            request.form['operator'],
            request.form['inspector'],
            request.form['production_qty']
        ])
        wb.save(file_path)
        return render_template('defects.html', success="Record saved.")
    return render_template('defects.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
